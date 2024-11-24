import os
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilenames
from openpyxl import load_workbook
from pyxlsb import open_workbook as open_xlsb
import random

def get_row_count(file_path):
    """
    统计文件行数的函数，支持 CSV 和 Excel 文件
    """
    try:
        if file_path.endswith(".csv"):
            # 尝试多种编码读取 CSV 文件，并只读取第一列以减少内存占用
            encodings = ['utf-8', 'gbk']
            for encoding in encodings:
                try:
                    with pd.read_csv(file_path, encoding=encoding, usecols=[0], chunksize=1000) as reader:
                        row_count = sum(len(chunk) for chunk in reader)
                    return row_count
                except Exception:
                    continue
            raise ValueError("无法读取 CSV 文件，可能编码不支持")

        elif file_path.endswith(".xlsx"):
            # 使用 openpyxl 快速统计行数（不加载整个文件）
            wb = load_workbook(filename=file_path, read_only=True)
            sheet = wb.active  # 默认读取第一个工作表
            return sheet.max_row  # 使用 sheet.max_row 获取行数

        elif file_path.endswith(".xlsb"):
            # 使用 pyxlsb 统计 .xlsb 文件行数
            with open_xlsb(file_path) as wb:
                with wb.get_sheet(1) as sheet:
                    return sum(1 for _ in sheet)  # 按行迭代统计行数

        elif file_path.endswith(".xls"):
            # 使用 pandas 读取 .xls 文件
            df = pd.read_excel(file_path, engine='xlrd', usecols=[0])
            return len(df)

        else:
            raise ValueError("不支持的文件格式")
    except Exception as e:
        print(f"读取文件 {file_path} 出错：{e}")
        return "读取失败"

def append_to_existing_file(existing_path, new_data):
    """
    将新的统计数据追加到现有的 Excel 文件的最后一列。
    """
    try:
        # 读取现有文件
        existing_df = pd.read_excel(existing_path, engine="openpyxl")

        # 如果新数据行数与已有数据行数不一致，则调整新数据的长度
        if len(existing_df) != len(new_data):
            max_len = max(len(existing_df), len(new_data))
            # 调整现有数据的行数
            existing_df = existing_df.reindex(range(max_len))
            # 调整新数据的行数
            new_data = new_data.reindex(range(max_len))

        # 将新数据按列追加到现有数据
        for column in new_data.columns:
            new_col_name = column
            while new_col_name in existing_df.columns:
                # 如果列名重复，在列名后追加数字避免冲突
                new_col_name = f"{column}_{random.randint(1, 1000)}"
            existing_df[new_col_name] = new_data[column].values

    except Exception as e:
        print(f"读取或合并现有文件出错：{e}")
        # 如果读取失败，直接使用新数据（按列模式）
        existing_df = new_data

    # 保存更新后的结果
    existing_df.to_excel(existing_path, index=False)
    print(f"结果已追加保存到：{existing_path}")


def count_file_rows():
    """
    主函数：选择文件并统计行数
    """
    # 打开文件选择对话框
    Tk().withdraw()  # 隐藏 Tkinter 主窗口
    file_paths = askopenfilenames(filetypes=[("Excel or CSV Files", "*.xls *.xlsx *.xlsb *.csv")], title="选择文件")

    if not file_paths:
        print("未选择任何文件")
        return

    stats = []  # 用于存储统计结果

    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        row_count = get_row_count(file_path)
        stats.append({"文件名": file_name, "行数": row_count})
        print(f"文件: {file_name}, 行数: {row_count}")

    # 将统计结果转换为 DataFrame
    stats_df = pd.DataFrame(stats)

    # 保存或追加统计结果
    save_dir = r"C:\excel"
    os.makedirs(save_dir, exist_ok=True)
    save_path = os.path.join(save_dir, "行数统计结果.xlsx")

    if os.path.exists(save_path):
        # 如果文件已存在，追加保存
        append_to_existing_file(save_path, stats_df)
    else:
        # 如果文件不存在，直接保存
        stats_df.to_excel(save_path, index=False)
        print(f"结果已保存到: {save_path}")

# if __name__ == "__main__":
#     count_file_rows()

def main():
    """
    统一入口，调用 process_files()，用于外部调用
    """
    count_file_rows()

if __name__ == "__main__":
    main()